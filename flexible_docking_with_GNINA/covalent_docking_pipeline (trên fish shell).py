# ============================================================
# 🧬 GNINA Covalent Flexible Docking Pipeline v2.7.0
#    Metalloprotein Mode (MMP-2 / Zn²⁺) — Linux + Fish Shell
# ============================================================
#
# ╔══════════════════════════════════════════════════════════════╗
# ║  CHANGELOG v2.6.2 → v2.7.0                                 ║
# ╠══════════════════════════════════════════════════════════════╣
# ║                                                              ║
# ║  🔴 FIX 1 — SCORING FUNCTION:                               ║
# ║     XÓA: --cnn_scoring rescore, --cnn_empirical_weight 1.0  ║
# ║     THÊM: --cnn_scoring none                                ║
# ║     Lý do: CNN của GNINA được huấn luyện trên PDBbind —     ║
# ║     tập dữ liệu chủ yếu non-covalent. Mô hình CHƯA học    ║
# ║     được coordination bond geometry (Zn²⁺-carboxylate).     ║
# ║     CNN scores sẽ trả về giá trị vô nghĩa hoặc penalty     ║
# ║     SAI cho các pose đúng → ranking bị hỏng.                ║
# ║     Ref: GNINA docs recommend --cnn_scoring none cho        ║
# ║     covalent docking.                                        ║
# ║                                                              ║
# ║  🔴 FIX 2 — POSE SORT ORDER:                                ║
# ║     XÓA: --pose_sort_order Energy (viết hoa — GNINA có thể  ║
# ║          không nhận diện)                                    ║
# ║     THÊM: --pose_sort_order energy (lowercase chuẩn)        ║
# ║     Khi CNN off, "energy" = xếp theo minimizedAffinity.     ║
# ║                                                              ║
# ║  🔴 FIX 3 — SCORING ENGINE:                                 ║
# ║     KHÔNG khai báo --scoring vinardo.                        ║
# ║     Giữ GNINA default = Vina scoring function.              ║
# ║     Lý do: Vinardo có Van der Waals radius cực kỳ khắt khe, ║
# ║     phạt steric clashes nặng hơn Vina gấp nhiều lần.        ║
# ║     Ligand bị neo tọa độ vào Zn²⁺ → clash ở sampling đầu   ║
# ║     → Vinardo đánh giá dương vô cực → UFF minimizer sụp đổ. ║
# ║                                                              ║
# ║  🟢 GIỮ NGUYÊN: Tất cả --covalent_* flags (đã có từ v2.6.2)║
# ║  🟢 GIỮ NGUYÊN: classify_red_flag, split_ligands,          ║
# ║     status management, prepare_root_folders                  ║
# ║                                                              ║
# ║  🔵 CẬP NHẬT OUTPUT:                                        ║
# ║     - Ranking: minimizedAffinity (thay vì CNN_VS)           ║
# ║     - Excel/CSV: CNN columns hiển thị "N/A (CNN off)"       ║
# ║     - Statistics: thêm Covalent Docking Parameters           ║
# ║     - Banner: ghi rõ Covalent Mode + CNN off                ║
# ║                                                              ║
# ╠══════════════════════════════════════════════════════════════╣
# ║  SMART COMPLIANCE                                            ║
# ║  S: Covalent docking MMP-2/Zn²⁺ với GNINA, CNN off         ║
# ║  M: minimizedAffinity (kcal/mol), sanity flags, pose count  ║
# ║  A: Single-GPU, exhaustiveness=64, timeout configurable     ║
# ║  R: Metalloprotein-specific — Vina empirical only           ║
# ║  T: Per-ligand timeout + total elapsed tracking             ║
# ║                                                              ║
# ║  FAIR COMPLIANCE                                             ║
# ║  F: Unique LIG_XXXX IDs, ligand_mapping.csv, STATUS.txt    ║
# ║  A: Standard SDF/PDB/CSV/XLSX, CLI reproducible             ║
# ║  I: SMILES, SDF V2000, RDKit-compatible                     ║
# ║  R: META.txt per ligand, command.txt, full provenance       ║
# ╚══════════════════════════════════════════════════════════════╝
#
# ============================================================

import os
import re
import subprocess
import time
import csv
import shutil
import socket
import traceback
from pathlib import Path
from rdkit import Chem

# =============================================================
# .env loading + LD_LIBRARY_PATH đảm bảo
# =============================================================
NOTEBOOK_DIR = Path.cwd()

_conda_prefix = os.environ.get("CONDA_PREFIX", "")
if _conda_prefix:
    _conda_lib = os.path.join(_conda_prefix, "lib")
    _current_ld = os.environ.get("LD_LIBRARY_PATH", "")
    if _conda_lib not in _current_ld:
        os.environ["LD_LIBRARY_PATH"] = (
            f"{_conda_lib}:{_current_ld}" if _current_ld else _conda_lib
        )
        print(f"🔧 Set LD_LIBRARY_PATH = {os.environ['LD_LIBRARY_PATH']}")

try:
    from dotenv import load_dotenv
    _env_file = NOTEBOOK_DIR / ".env"
    if _env_file.exists():
        load_dotenv(_env_file, override=True)
        print(f"📄 Loaded .env from {_env_file}")
except ImportError:
    print("⚠️ python-dotenv not installed. Using environment variables only.")


def _resolve_path(env_var: str, fallback: str) -> str:
    raw = os.environ.get(env_var, fallback)
    expanded = os.path.expanduser(raw)
    if not os.path.isabs(expanded):
        expanded = str(NOTEBOOK_DIR / expanded)
    return os.path.normpath(expanded)


# =========================
# GLOBAL CONFIG
# =========================
BASE_DIR = _resolve_path("DOCKING_BASE_DIR", str(NOTEBOOK_DIR))
RESULTS_DIR = f"{BASE_DIR}/docking_results"
PROTEIN_PATH = _resolve_path(
    "PROTEIN_PATH",
    f"{BASE_DIR}/data/mmp2_7xjo_ready_for_gnina.pdb",
)
REF_LIGAND = _resolve_path("REF_LIGAND", f"{BASE_DIR}/data/ref_ligand.sdf")
LIGAND_SDF = _resolve_path("LIGAND_SDF", f"{BASE_DIR}/data/ligands_prepared.sdf")
FLEX_RESIDUES = "A:180,A:181,A:215,A:235,A:240"
SEED = "42"
GPU_DEVICE = os.environ.get("GNINA_GPU_DEVICE", "0")

# ================================================================
# Arbitration thresholds — Thermodynamic Sanity Only
# ================================================================
# [v2.7.0] CNN_VS: KHÔNG có ngưỡng (CNN đã tắt cho covalent mode).
#
# Red Flag chỉ dựa trên thermodynamic sanity:
#   🟡 POSITIVE_AFFINITY: affinity > 0 (repulsive, vô nghĩa vật lý)
#   🟠 POOR_AFFINITY: affinity ≥ threshold (gắn kết quá yếu)
# ================================================================
RED_FLAG_AFFINITY_POOR = -6.5

# =====================================================
# GNINA timeout — configurable via .env
# =====================================================
GNINA_TIMEOUT_SEC = int(os.environ.get("GNINA_TIMEOUT_SEC", "3600"))

# GNINA_BIN — auto-detect
_env_bin = os.environ.get("GNINA_BIN", "")
if _env_bin and os.path.isfile(_env_bin):
    GNINA_BIN = _env_bin
elif shutil.which("gnina"):
    GNINA_BIN = shutil.which("gnina")
else:
    GNINA_BIN = f"{BASE_DIR}/bin/gnina"

# Status constants
STATUS_PENDING = "PENDING"
STATUS_RUNNING = "RUNNING"
STATUS_DONE = "DONE"
STATUS_FAILED = "FAILED"


# =========================
# Subprocess env
# =========================
def _get_subprocess_env() -> dict:
    env = os.environ.copy()

    conda_prefix = env.get("CONDA_PREFIX", "")
    if conda_prefix:
        conda_lib = os.path.join(conda_prefix, "lib")
        current_ld = env.get("LD_LIBRARY_PATH", "")
        if conda_lib not in current_ld:
            env["LD_LIBRARY_PATH"] = (
                f"{conda_lib}:{current_ld}" if current_ld else conda_lib
            )

    cuda_vis = os.environ.get("CUDA_VISIBLE_DEVICES")
    if cuda_vis is not None:
        env["CUDA_VISIBLE_DEVICES"] = cuda_vis

    return env


# =========================
# STATUS MANAGEMENT
# =========================
def write_status(lig_root: str, status: str, **kwargs):
    status_file = os.path.join(lig_root, "STATUS.txt")

    if status == STATUS_RUNNING:
        with open(status_file, "w") as f:
            f.write(f"STATUS={status}\n")
            f.write(f"HOST={socket.gethostname()}\n")
            f.write(f"START_TIME={time.strftime('%Y-%m-%d %H:%M:%S')}\n")
    else:
        with open(status_file, "a") as f:
            f.write(f"STATUS={status}\n")
            for key, value in kwargs.items():
                f.write(f"{key.upper()}={value}\n")
            f.write(f"END_TIME={time.strftime('%Y-%m-%d %H:%M:%S')}\n")


def read_status(lig_root: str) -> str:
    status_file = os.path.join(lig_root, "STATUS.txt")

    if not os.path.exists(status_file):
        return STATUS_PENDING

    last_status = STATUS_PENDING
    try:
        with open(status_file, "r") as f:
            for line in f:
                if line.startswith("STATUS="):
                    last_status = line.strip().split("=", 1)[1]
    except Exception:
        pass

    return last_status


def get_status_details(lig_root: str) -> dict:
    status_file = os.path.join(lig_root, "STATUS.txt")
    details = {}

    if os.path.exists(status_file):
        with open(status_file, "r") as f:
            for line in f:
                if "=" in line:
                    key, value = line.strip().split("=", 1)
                    details[key] = value

    return details


# =========================
# UTILS
# =========================
def sanitize_name(name: str, max_len: int = 80) -> str:
    if not name:
        return "NA"

    name = name.strip().replace(" ", "_")
    name = re.sub(r'[\/:*?"<>|\\]', "", name)
    name = re.sub(r"_+", "_", name)
    name = name.strip("_")

    return name[:max_len] if name else "NA"


# =========================
# PREPARE ROOT FOLDERS
# =========================
def prepare_root_folders():
    dirs = [
        f"{RESULTS_DIR}/protein",
        f"{RESULTS_DIR}/reference",
        f"{RESULTS_DIR}/ligands",
        f"{RESULTS_DIR}/summary",
    ]

    for d in dirs:
        os.makedirs(d, exist_ok=True)

    protein_dest = f"{RESULTS_DIR}/protein/receptor.pdb"
    ref_dest = f"{RESULTS_DIR}/reference/ref_ligand.sdf"

    if not os.path.exists(protein_dest):
        shutil.copy(PROTEIN_PATH, protein_dest)
        print(f"✔ Copied receptor to {protein_dest}")

    if not os.path.exists(ref_dest):
        shutil.copy(REF_LIGAND, ref_dest)
        print(f"✔ Copied reference ligand to {ref_dest}")


# =========================
# SPLIT LIGANDS
# =========================
def split_ligands(input_sdf: str, ligands_root: str) -> list:
    os.makedirs(ligands_root, exist_ok=True)

    cleaned_sdf = os.path.join(ligands_root, "_cleaned_input.sdf")
    with open(input_sdf, "rb") as f_in:
        raw = f_in.read()

    if b"\r" in raw:
        raw = raw.replace(b"\r\n", b"\n").replace(b"\r", b"\n")
        with open(cleaned_sdf, "wb") as f_out:
            f_out.write(raw)
        sdf_to_parse = cleaned_sdf
        print(f"🔧 Found \\r in input — normalized line endings")
    else:
        sdf_to_parse = input_sdf
        print(f"✔ Input SDF has clean line endings — no fix needed")

    suppl = Chem.SDMolSupplier(sdf_to_parse, removeHs=False, sanitize=True)
    ligands = []
    mapping = []
    forced_3d_list = []

    for idx, mol in enumerate(suppl, start=1):
        if mol is None:
            print(f"⚠️ Skipping invalid molecule at index {idx}")
            continue

        lig_id = f"LIG_{idx:04d}"

        conf = mol.GetConformer() if mol.GetNumConformers() > 0 else None
        if conf is not None:
            has_z = any(
                abs(conf.GetAtomPosition(i).z) > 0.001
                for i in range(mol.GetNumAtoms())
            )
            if has_z and not conf.Is3D():
                conf.Set3D(True)
                forced_3d_list.append(lig_id)
                print(f"🔧 {lig_id}: forced 3D tag (had Z coordinates)")

        orig_name = mol.GetProp("_Name") if mol.HasProp("_Name") else "NA"
        safe_name = sanitize_name(orig_name)

        lig_dirname = f"{lig_id}__{safe_name}"
        lig_root = os.path.join(ligands_root, lig_dirname)
        input_dir = os.path.join(lig_root, "input")
        os.makedirs(input_dir, exist_ok=True)

        ligand_sdf = os.path.join(input_dir, "ligand.sdf")
        writer = Chem.SDWriter(ligand_sdf)
        writer.SetForceV3000(False)
        writer.write(mol)
        writer.close()

        smiles = Chem.MolToSmiles(mol)
        with open(os.path.join(lig_root, "META.txt"), "w") as f:
            f.write(f"ID={lig_id}\n")
            f.write(f"DIR_NAME={lig_dirname}\n")
            f.write(f"ORIGINAL_NAME={orig_name}\n")
            f.write(f"SMILES={smiles}\n")
            f.write(f"SDF_INDEX={idx}\n")

        ligands.append(
            {
                "lig_id": lig_id,
                "lig_dirname": lig_dirname,
                "lig_root": lig_root,
                "ligand_sdf": ligand_sdf,
                "orig_name": orig_name,
                "smiles": smiles,
            }
        )
        mapping.append((lig_id, lig_dirname, orig_name, smiles))

    mapping_file = os.path.join(ligands_root, "ligand_mapping.csv")
    with open(mapping_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "DIR_NAME", "ORIGINAL_NAME", "SMILES"])
        for row in mapping:
            writer.writerow(row)

    if forced_3d_list:
        forced_log = os.path.join(ligands_root, "forced_3d_log.txt")
        with open(forced_log, "w") as f:
            f.write("# Ligands forced from 2D tag to 3D tag\n")
            f.write(f"# Total: {len(forced_3d_list)}\n")
            for lid in forced_3d_list:
                f.write(f"{lid}\n")
        print(f"📝 Forced 3D log: {forced_log} ({len(forced_3d_list)} ligands)")

    count_3d = 0
    count_2d = 0
    for lig in ligands:
        mol = next(Chem.SDMolSupplier(lig["ligand_sdf"], removeHs=False))
        if mol is not None and mol.GetNumConformers() > 0:
            if mol.GetConformer().Is3D():
                count_3d += 1
            else:
                count_2d += 1
    print(f"✔ Validation: {count_3d} molecules 3D, {count_2d} still 2D")
    if count_2d > 0:
        print(f"⚠️ WARNING: {count_2d} molecules still tagged 2D — check input geometry")

    if os.path.exists(cleaned_sdf):
        os.remove(cleaned_sdf)

    print(f"✔ Split {len(ligands)} ligands")
    print(f"✔ Mapping written to {mapping_file}")

    return ligands


# ================================================================
# SCORE PARSING
# ================================================================
def parse_top_poses(sdf_path: str, n: int = 3) -> list:
    """
    Extract top N poses from docked SDF.

    [v2.7.0] Khi --cnn_scoring none:
      - CNNscore, CNNaffinity, CNN_VS sẽ KHÔNG có trong SDF output
      - Hàm trả về None cho các trường này (không crash)
      - minimizedAffinity là metric chính duy nhất
      - Thứ tự poses = thứ tự energy từ GNINA (đã sort bởi
        --pose_sort_order energy)
    """
    try:
        suppl = Chem.SDMolSupplier(sdf_path, removeHs=False)
        poses = []

        for pose_idx, mol in enumerate(suppl, start=1):
            if mol is None:
                continue

            pose_data = {"pose_rank": pose_idx}

            score_props = [
                "minimizedAffinity",
                "CNNscore",
                "CNNaffinity",
                "CNN_VS",
                "vinardo",
            ]

            for prop in score_props:
                if mol.HasProp(prop):
                    try:
                        pose_data[prop] = float(mol.GetProp(prop))
                    except ValueError:
                        pose_data[prop] = mol.GetProp(prop)
                else:
                    pose_data[prop] = None

            poses.append(pose_data)

            if len(poses) >= n:
                break

        return poses

    except Exception as e:
        print(f"⚠️ Error parsing scores: {e}")
        return []


# ================================================================
# RED FLAG CLASSIFICATION — Thermodynamic Sanity Only
# ================================================================
def classify_red_flag(affinity) -> str:
    """
    Arbitration Protocol — Tầng 3: Thermodynamic Sanity Check.

    [v2.7.0] Chỉ kiểm tra affinity. CNN đã tắt hoàn toàn.

    Sanity flags:
      🟡 POSITIVE_AFFINITY: affinity > 0 kcal/mol (repulsive)
      🟠 POOR_AFFINITY: affinity ≥ threshold (gắn kết quá yếu)
    """
    if affinity is None:
        return ""

    if affinity > 0:
        return "🟡 POSITIVE_AFFINITY"

    if affinity >= RED_FLAG_AFFINITY_POOR:
        return "🟠 POOR_AFFINITY"

    return ""


def get_best_pose_summary(sdf_path: str) -> dict:
    """
    Lấy pose tốt nhất (pose 1 = energy thấp nhất) và tính red flag.
    """
    poses = parse_top_poses(sdf_path, n=1)
    if not poses:
        return {
            "CNNscore": None,
            "CNN_VS": None,
            "CNNaffinity": None,
            "minimizedAffinity": None,
            "red_flag": "",
        }

    best = poses[0]
    best["red_flag"] = classify_red_flag(best.get("minimizedAffinity"))
    return best


def parse_best_score(sdf_path: str):
    """Extract best minimizedAffinity from docked SDF.

    Returns None nếu không parse được.
    """
    poses = parse_top_poses(sdf_path, n=1)
    if poses and poses[0].get("minimizedAffinity") is not None:
        return poses[0]["minimizedAffinity"]
    return None


# ================================================================
# RUN GNINA — Covalent Metalloprotein Mode
# ================================================================
def run_gnina(ligand_info: dict, idx: int, total: int) -> bool:
    """Run GNINA covalent docking for a single ligand.

    [v2.7.0] Covalent metalloprotein configuration:
      - --cnn_scoring none (CNN chưa calibrate cho coordination bond)
      - --pose_sort_order energy (xếp theo minimizedAffinity)
      - Scoring function: Vina default (KHÔNG dùng Vinardo)
      - Tất cả --covalent_* flags: GIỮ NGUYÊN
      - timeout=GNINA_TIMEOUT_SEC (configurable)
    """
    lig_id = ligand_info["lig_id"]
    lig_root = ligand_info["lig_root"]
    ligand_sdf = ligand_info["ligand_sdf"]

    out_dir = os.path.join(lig_root, "output")
    log_dir = os.path.join(lig_root, "logs")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)

    out_lig = os.path.join(out_dir, "docked.sdf")
    out_flex = os.path.join(out_dir, "flex_residues.pdb")
    log_file = os.path.join(log_dir, "gnina.log")
    stderr_file = os.path.join(log_dir, "gnina_stderr.log")
    cmd_file = os.path.join(log_dir, "command.txt")

    write_status(lig_root, STATUS_RUNNING)

    print(f"\n🔄 [{idx}/{total}] Docking {lig_id} ...")
    start = time.time()

    # ══════════════════════════════════════════════════════════
    # GNINA COMMAND — v2.7.0 Covalent Metalloprotein
    # ══════════════════════════════════════════════════════════
    cmd = [
        GNINA_BIN,
        # ── Receptor & Ligand ──
        "-r",
        f"{RESULTS_DIR}/protein/receptor.pdb",
        "-l",
        ligand_sdf,
        # ── Search space (autobox from reference ligand) ──
        "--autobox_ligand",
        f"{RESULTS_DIR}/reference/ref_ligand.sdf",
        "--autobox_add",
        "5",
        "--autobox_extend",
        "1",
        # ── Flexible residues ──
        "--flexres",
        FLEX_RESIDUES,
        # ── Sampling parameters ──
        "--num_modes",
        "10",
        "--exhaustiveness",
        "64",
        # ──────────────────────────────────────────────────────
        # COVALENT ANCHORING BLOCK
        # Target: MMP-2 catalytic Zn²⁺ (residue A:301)
        # Ligand anchor: carboxylate oxygen [OX1;$([O]C=O)]
        # Coordinates: from PLIP analysis of PDB 7XJO
        #
        # --covalent_rec_atom: Chỉ định nguyên tử receptor làm
        #   mỏ neo (Zn²⁺ trong chuỗi A, residue 301)
        # --covalent_lig_atom_pattern: SMARTS pattern xác định
        #   nguyên tử ligand sẽ gắn vào mỏ neo
        # --covalent_lig_atom_position: Tọa độ 3D (Å) mà
        #   nguyên tử ligand phải nằm tại
        # --covalent_fix_lig_atom_position: Khóa cứng tọa độ
        #   nguyên tử neo trong suốt quá trình sampling
        # --covalent_optimize_lig: Cho phép tối ưu phần còn
        #   lại của ligand sau khi neo
        # ──────────────────────────────────────────────────────
        "--covalent_rec_atom",
        "A:301:ZN",
        "--covalent_lig_atom_pattern",
        "[OX1;$([O]C=O)]",
        "--covalent_lig_atom_position",
        "6.739,10.721,31.893",
        "--covalent_fix_lig_atom_position",
        "--covalent_optimize_lig",
        # ──────────────────────────────────────────────────────
        # SCORING CONFIGURATION — v2.7.0
        #
        # --cnn_scoring none:
        #   CNN của GNINA (CrossDock2020/PDBbind training set)
        #   CHƯA được huấn luyện cho:
        #     ① Coordination bonds (Zn²⁺-O, Zn²⁺-N)
        #     ② Covalent attachment geometry
        #     ③ Metal-mediated interactions
        #   Khi bật CNN cho covalent metalloprotein docking:
        #     - CNNscore trả về giá trị KHÔNG ĐÁNG TIN CẬY
        #     - CNN có thể penalty pose đúng vì geometry "lạ"
        #     - Pose ranking bị nhiễu → top poses sai
        #   → Tắt CNN, chỉ dùng Vina empirical scoring.
        #
        # KHÔNG khai báo --scoring vinardo:
        #   Vinardo sử dụng Van der Waals radius cực kỳ khắt
        #   khe. Khi ligand bị neo tọa độ vào Zn²⁺:
        #     - Sampling đầu CHẮC CHẮN có steric clashes
        #     - Vinardo phạt → energy dương vô cực
        #     - UFF minimizer nhận input quá lớn → sụp đổ
        #   Vina default bao dung hơn → minimizer hội tụ.
        #
        # --pose_sort_order energy:
        #   Khi CNN off, không có CNNscore để sort.
        #   "energy" = sort theo minimizedAffinity:
        #     - Giá trị càng âm = gắn kết càng mạnh
        #     - Pose 1 = pose năng lượng thấp nhất (tốt nhất)
        # ──────────────────────────────────────────────────────
        "--cnn_scoring",
        "none",
        "--pose_sort_order",
        "energy",
        # ── Hardware ──
        "--device",
        GPU_DEVICE,
        "--seed",
        SEED,
        # ── Output ──
        "--atom_term_data",
        "-o",
        out_lig,
        "--out_flex",
        out_flex,
        "--log",
        log_file,
    ]

    # ── FAIR: Lưu command để reproducibility ──
    with open(cmd_file, "w") as f:
        f.write(" \\\n    ".join(cmd))

    try:
        with open(stderr_file, "w") as stderr_f:
            subprocess.run(
                cmd,
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=stderr_f,
                text=True,
                env=_get_subprocess_env(),
                timeout=GNINA_TIMEOUT_SEC,
            )

        if not os.path.exists(out_lig) or os.path.getsize(out_lig) == 0:
            raise RuntimeError("GNINA produced empty output SDF")

        elapsed = (time.time() - start) / 60

        best_score = parse_best_score(out_lig)

        write_status(
            lig_root,
            STATUS_DONE,
            elapsed_min=f"{elapsed:.2f}",
            best_affinity=(
                f"{best_score:.4f}" if best_score is not None else "NA"
            ),
        )

        score_str = f"{best_score:.4f}" if best_score is not None else "N/A"
        print(
            f"✅ [{idx}/{total}] {lig_id} DONE in {elapsed:.2f} min"
            f" (affinity: {score_str})"
        )
        return True

    except subprocess.TimeoutExpired:
        elapsed = (time.time() - start) / 60
        timeout_min = GNINA_TIMEOUT_SEC / 60
        write_status(
            lig_root,
            STATUS_FAILED,
            elapsed_min=f"{elapsed:.2f}",
            error=f"TIMEOUT after {GNINA_TIMEOUT_SEC}s ({timeout_min:.0f}min)",
        )
        print(
            f"⏰ [{idx}/{total}] {lig_id} TIMEOUT after"
            f" {timeout_min:.0f} min — skipping"
        )
        return False

    except subprocess.CalledProcessError as e:
        elapsed = (time.time() - start) / 60
        write_status(
            lig_root,
            STATUS_FAILED,
            elapsed_min=f"{elapsed:.2f}",
            error=f"GNINA exit code {e.returncode}",
        )
        print(
            f"❌ [{idx}/{total}] {lig_id} FAILED: GNINA exit code {e.returncode}"
        )
        return False

    except Exception as e:
        elapsed = (time.time() - start) / 60
        write_status(
            lig_root,
            STATUS_FAILED,
            elapsed_min=f"{elapsed:.2f}",
            error=str(e),
            traceback=traceback.format_exc().replace("\n", " | "),
        )
        print(f"❌ [{idx}/{total}] {lig_id} FAILED: {e}")
        return False


# ================================================================
# EXCEL / CSV SUMMARY — v2.7.0 Covalent Mode
# ================================================================
def generate_excel_summary(ligands: list, summary_dir: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        USE_OPENPYXL = True
    except ImportError:
        print("⚠️ openpyxl not installed, generating CSV instead")
        USE_OPENPYXL = False

    results = []
    for lig in ligands:
        lig_root = lig["lig_root"]
        out_sdf = os.path.join(lig_root, "output", "docked.sdf")

        status = read_status(lig_root)
        details = get_status_details(lig_root)

        top_poses = []
        best_pose = {}
        if status == STATUS_DONE and os.path.exists(out_sdf):
            top_poses = parse_top_poses(out_sdf, n=3)
            best_pose = get_best_pose_summary(out_sdf)

        results.append(
            {
                "lig_id": lig["lig_id"],
                "lig_dirname": lig["lig_dirname"],
                "orig_name": lig["orig_name"],
                "smiles": lig["smiles"],
                "status": status,
                "elapsed_min": details.get("ELAPSED_MIN", ""),
                "top_poses": top_poses,
                "best_CNN_VS": best_pose.get("CNN_VS"),
                "best_CNNscore": best_pose.get("CNNscore"),
                "best_CNNaffinity": best_pose.get("CNNaffinity"),
                "best_affinity": best_pose.get("minimizedAffinity"),
                "red_flag": best_pose.get("red_flag", ""),
            }
        )

    # ──────────────────────────────────────────────────────────────
    # [v2.7.0] PRIMARY RANKING: minimizedAffinity (càng âm càng tốt)
    #
    # Khi CNN tắt, CNN_VS = None cho mọi ligand → không thể rank
    # theo CNN_VS. Rank theo minimizedAffinity:
    #   - Giá trị càng âm = gắn kết càng mạnh = rank cao hơn
    #   - None values xếp cuối
    # ──────────────────────────────────────────────────────────────
    results_ranked = sorted(
        results,
        key=lambda x: (
            x["best_affinity"] is None,
            x["best_affinity"] if x["best_affinity"] is not None else 0,
        ),
    )

    if USE_OPENPYXL:
        _generate_xlsx_v27(results, results_ranked, summary_dir)
    else:
        _generate_csv_fallback_v27(results_ranked, summary_dir)


def _generate_xlsx_v27(
    results: list, results_ranked: list, summary_dir: str
):
    """
    [v2.7.0] Excel output cho Covalent Metalloprotein Docking (CNN off).

    4 Sheets:
      Sheet 1 — Intra-Ligand Poses: Top 3 poses mỗi ligand (energy sort)
      Sheet 2 — Inter-Ligand Ranking: Ranking theo minimizedAffinity
      Sheet 3 — Sanity Flags: Ligands vi phạm thermodynamic sanity
      Sheet 4 — Statistics: Pipeline metadata + affinity distribution
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Shared styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    header_fill_green = PatternFill(
        start_color="548235", end_color="548235", fill_type="solid"
    )
    header_fill_red = PatternFill(
        start_color="C00000", end_color="C00000", fill_type="solid"
    )
    header_fill_purple = PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    )
    done_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    failed_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    caution_fill = PatternFill(
        start_color="FFD93D", end_color="FFD93D", fill_type="solid"
    )
    poor_affinity_fill = PatternFill(
        start_color="FFA94D", end_color="FFA94D", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def _write_header(ws, row, headers, fill):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    def _set_col_widths(ws, widths: dict):
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    def _fmt_cnn(val):
        """Format CNN value — hiển thị 'N/A (CNN off)' khi None."""
        if val is None:
            return "N/A (CNN off)"
        return val

    # ══════════════════════════════════════════════════════════════
    # SHEET 1: Intra-Ligand Poses (Tầng 1)
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Intra-Ligand_Poses"

    ws1.merge_cells("A1:S1")
    note_cell = ws1.cell(
        row=1,
        column=1,
        value=(
            "TẦNG 1 — INTRA-LIGAND POSE SELECTION (Covalent Mode): "
            "Poses xếp theo energy (minimizedAffinity) vì CNN đã tắt "
            "(--cnn_scoring none cho metalloprotein). "
            "Pose 1 = pose có năng lượng gắn kết thấp nhất (tốt nhất). "
            "Các cột CNN hiển thị N/A — không tính khi CNN off."
        ),
    )
    note_cell.font = Font(bold=True, italic=True, size=10, color="2F5496")

    headers_s1 = [
        "#",
        "ID",
        "Original_Name",
        "Status",
        "Time_min",
        "P1_Affinity",
        "P1_CNNscore",
        "P1_CNN_VS",
        "P1_CNNaffinity",
        "P1_Flag",
        "P2_Affinity",
        "P2_CNNscore",
        "P2_CNN_VS",
        "P2_CNNaffinity",
        "P3_Affinity",
        "P3_CNNscore",
        "P3_CNN_VS",
        "P3_CNNaffinity",
        "SMILES",
    ]

    _write_header(ws1, 2, headers_s1, header_fill_blue)
    ws1.freeze_panes = "A3"

    for row_idx, data in enumerate(results_ranked, start=3):
        rank = row_idx - 2
        top_poses = data["top_poses"]

        row_data = [
            rank,
            data["lig_id"],
            data["orig_name"],
            data["status"],
            data["elapsed_min"],
        ]

        # Pose 1
        if len(top_poses) >= 1:
            p = top_poses[0]
            flag = classify_red_flag(p.get("minimizedAffinity"))
            row_data.extend([
                p.get("minimizedAffinity", ""),
                _fmt_cnn(p.get("CNNscore")),
                _fmt_cnn(p.get("CNN_VS")),
                _fmt_cnn(p.get("CNNaffinity")),
                flag,
            ])
        else:
            row_data.extend(["", "", "", "", ""])

        # Pose 2
        if len(top_poses) >= 2:
            p = top_poses[1]
            row_data.extend([
                p.get("minimizedAffinity", ""),
                _fmt_cnn(p.get("CNNscore")),
                _fmt_cnn(p.get("CNN_VS")),
                _fmt_cnn(p.get("CNNaffinity")),
            ])
        else:
            row_data.extend(["", "", "", ""])

        # Pose 3
        if len(top_poses) >= 3:
            p = top_poses[2]
            row_data.extend([
                p.get("minimizedAffinity", ""),
                _fmt_cnn(p.get("CNNscore")),
                _fmt_cnn(p.get("CNN_VS")),
                _fmt_cnn(p.get("CNNaffinity")),
            ])
        else:
            row_data.extend(["", "", "", ""])

        row_data.append(data["smiles"])

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_idx == 4:
                if value == STATUS_DONE:
                    cell.fill = done_fill
                elif value == STATUS_FAILED:
                    cell.fill = failed_fill

            if col_idx == 10:
                if "POSITIVE" in str(value):
                    cell.fill = caution_fill
                    cell.font = Font(bold=True)
                elif "POOR" in str(value):
                    cell.fill = poor_affinity_fill
                    cell.font = Font(bold=True)

    _set_col_widths(ws1, {
        "A": 5, "B": 12, "C": 35, "D": 10, "E": 9,
        "F": 11, "G": 15, "H": 15, "I": 15, "J": 20,
        "K": 11, "L": 15, "M": 15, "N": 15,
        "O": 11, "P": 15, "Q": 15, "R": 15,
        "S": 60,
    })

    # ══════════════════════════════════════════════════════════════
    # SHEET 2: Inter-Ligand Ranking (Tầng 2)
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Inter-Ligand_Ranking")

    ws2.merge_cells("A1:I1")
    note2 = ws2.cell(
        row=1,
        column=1,
        value=(
            "TẦNG 2 — INTER-LIGAND RANKING (Covalent Mode — CNN off): "
            "Xếp hạng theo minimizedAffinity (kcal/mol, càng âm càng tốt). "
            "CNN_VS không khả dụng vì --cnn_scoring none cho metalloprotein. "
            "Affinity là proxy duy nhất cho binding strength. "
            "Cột Flag = Tầng 3 Thermodynamic Sanity Check."
        ),
    )
    note2.font = Font(bold=True, italic=True, size=10, color="548235")

    headers_s2 = [
        "Energy_Rank",
        "ID",
        "Original_Name",
        "Affinity_kcal",
        "CNNscore",
        "CNN_VS",
        "CNNaffinity",
        "Sanity_Flag",
        "SMILES",
    ]

    _write_header(ws2, 2, headers_s2, header_fill_green)
    ws2.freeze_panes = "A3"

    top_done = [r for r in results_ranked if r["status"] == STATUS_DONE][:50]

    for row_idx, data in enumerate(top_done, start=3):
        rank = row_idx - 2

        row_data = [
            rank,
            data["lig_id"],
            data["orig_name"],
            data["best_affinity"],
            _fmt_cnn(data["best_CNNscore"]),
            _fmt_cnn(data["best_CNN_VS"]),
            _fmt_cnn(data["best_CNNaffinity"]),
            data["red_flag"],
            data["smiles"],
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_idx == 4 and isinstance(value, (int, float)):
                if value > 0:
                    cell.fill = caution_fill
                    cell.font = Font(bold=True, color="C00000")
                elif value >= RED_FLAG_AFFINITY_POOR:
                    cell.fill = PatternFill(
                        start_color="FFF2CC",
                        end_color="FFF2CC",
                        fill_type="solid",
                    )

            if col_idx == 8:
                if "POSITIVE" in str(value):
                    cell.fill = caution_fill
                    cell.font = Font(bold=True)
                elif "POOR" in str(value):
                    cell.fill = poor_affinity_fill
                    cell.font = Font(bold=True)

    _set_col_widths(ws2, {
        "A": 12, "B": 12, "C": 35, "D": 13, "E": 15,
        "F": 15, "G": 15, "H": 22, "I": 60,
    })

    # ══════════════════════════════════════════════════════════════
    # SHEET 3: Sanity Flags (Tầng 3)
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet(title="Sanity_Flags")

    ws3.merge_cells("A1:J1")
    note3 = ws3.cell(
        row=1,
        column=1,
        value=(
            "TẦNG 3 — THERMODYNAMIC SANITY CHECK: "
            "🟡 POSITIVE_AFFINITY = Affinity > 0 kcal/mol (repulsive, "
            "vô nghĩa vật lý). "
            f"🟠 POOR_AFFINITY = Affinity ≥ {RED_FLAG_AFFINITY_POOR} "
            "kcal/mol (gắn kết quá yếu). "
            "Các ligand này CẦN kiểm tra thủ công bằng visualisation. "
            "CNN metrics N/A (covalent mode — CNN off)."
        ),
    )
    note3.font = Font(bold=True, italic=True, size=10, color="C00000")

    headers_s3 = [
        "#",
        "ID",
        "Original_Name",
        "Flag_Type",
        "Affinity_kcal",
        "CNNscore",
        "CNN_VS",
        "CNNaffinity",
        "Concern",
        "SMILES",
    ]

    _write_header(ws3, 2, headers_s3, header_fill_red)
    ws3.freeze_panes = "A3"

    flagged = [r for r in results_ranked if r["red_flag"]]

    if flagged:
        for row_idx, data in enumerate(flagged, start=3):
            flag = data["red_flag"]

            if "POSITIVE" in flag:
                concern = (
                    f"Affinity={data['best_affinity']:.2f} > 0 (repulsive). "
                    "Pose này vi phạm nhiệt động học cơ bản — "
                    "ligand bị đẩy ra khỏi túi gắn."
                )
            elif "POOR" in flag:
                concern = (
                    f"Affinity={data['best_affinity']:.2f} ≥ "
                    f"{RED_FLAG_AFFINITY_POOR} kcal/mol. "
                    "Gắn kết quá yếu — không đủ ý nghĩa sinh học. "
                    "Cần kiểm tra tương tác bằng visualisation."
                )
            else:
                concern = ""

            row_data = [
                row_idx - 2,
                data["lig_id"],
                data["orig_name"],
                flag,
                data["best_affinity"],
                _fmt_cnn(data["best_CNNscore"]),
                _fmt_cnn(data["best_CNN_VS"]),
                _fmt_cnn(data["best_CNNaffinity"]),
                concern,
                data["smiles"],
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                if col_idx == 4:
                    if "POSITIVE" in str(value):
                        cell.fill = caution_fill
                        cell.font = Font(bold=True)
                    elif "POOR" in str(value):
                        cell.fill = poor_affinity_fill
                        cell.font = Font(bold=True)
    else:
        ws3.cell(
            row=3,
            column=1,
            value="✅ Không có ligand nào bị gắn cờ — Tất cả đều pass sanity check.",
        ).font = Font(italic=True, color="548235", size=12)

    _set_col_widths(ws3, {
        "A": 5, "B": 12, "C": 35, "D": 22, "E": 13,
        "F": 15, "G": 15, "H": 15, "I": 60, "J": 60,
    })

    # ══════════════════════════════════════════════════════════════
    # SHEET 4: Statistics — Pipeline Metadata + Distributions
    # ══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet(title="Statistics")

    total = len(results)
    done = sum(1 for r in results if r["status"] == STATUS_DONE)
    failed = sum(1 for r in results if r["status"] == STATUS_FAILED)
    pending = sum(1 for r in results if r["status"] == STATUS_PENDING)

    done_results = [r for r in results if r["best_affinity"] is not None]
    affinities = [r["best_affinity"] for r in done_results]

    n_poor = sum(1 for r in results if "POOR" in r.get("red_flag", ""))
    n_positive = sum(
        1 for r in results if "POSITIVE" in r.get("red_flag", "")
    )
    n_flagged = n_poor + n_positive

    n_timeout = sum(
        1
        for r in results
        if r["status"] == STATUS_FAILED
        and "TIMEOUT" in get_status_details(
            os.path.join(
                RESULTS_DIR, "ligands", r["lig_dirname"]
            )
        ).get("ERROR", "")
    )

    stats = [
        ("═══ PIPELINE OVERVIEW ═══", ""),
        ("Pipeline Version", "v2.7.0 — Covalent Metalloprotein"),
        ("Docking Mode", "Covalent Flexible (--covalent_fix_lig_atom_position)"),
        ("Scoring Function", "Vina (GNINA default — CNN off)"),
        ("CNN Scoring", "DISABLED (--cnn_scoring none)"),
        (
            "Rationale",
            "CNN not trained on coordination bonds; "
            "scores unreliable for Zn²⁺ metalloprotein",
        ),
        ("Pose Sort Order", "energy (minimizedAffinity)"),
        ("Total Ligands", total),
        ("Completed", done),
        ("Failed", failed),
        ("  ↳ of which TIMEOUT", n_timeout),
        ("Pending", pending),
        (
            "GNINA Timeout Setting",
            f"{GNINA_TIMEOUT_SEC}s ({GNINA_TIMEOUT_SEC/60:.0f}min)",
        ),
        ("", ""),
        ("═══ COVALENT DOCKING PARAMETERS ═══", ""),
        ("Receptor Anchor Atom", "A:301:ZN"),
        ("Ligand Atom Pattern", "[OX1;$([O]C=O)]"),
        ("Anchor Position (Å)", "6.739, 10.721, 31.893"),
        ("Fix Ligand Atom Position", "YES"),
        ("Optimize Ligand", "YES"),
        ("Flexible Residues", FLEX_RESIDUES),
        ("Autobox Add (Å)", "5"),
        ("Autobox Extend", "1"),
        ("Exhaustiveness", "64"),
        ("Num Modes", "10"),
        ("Random Seed", SEED),
        ("", ""),
        ("═══ AFFINITY DISTRIBUTION (kcal/mol) ═══", ""),
        (
            "Best Affinity",
            f"{min(affinities):.4f}" if affinities else "N/A",
        ),
        (
            "Worst Affinity",
            f"{max(affinities):.4f}" if affinities else "N/A",
        ),
        (
            "Mean Affinity",
            (
                f"{sum(affinities) / len(affinities):.4f}"
                if affinities
                else "N/A"
            ),
        ),
        (
            "Median Affinity",
            (
                f"{sorted(affinities)[len(affinities)//2]:.4f}"
                if affinities
                else "N/A"
            ),
        ),
        (
            "Ligands Affinity < -10.0",
            sum(1 for a in affinities if a < -10.0),
        ),
        (
            "Ligands Affinity < -9.0",
            sum(1 for a in affinities if a < -9.0),
        ),
        (
            "Ligands Affinity < -8.0",
            sum(1 for a in affinities if a < -8.0),
        ),
        (
            "Ligands Affinity < -7.0",
            sum(1 for a in affinities if a < -7.0),
        ),
        (
            "Ligands Affinity < -6.0",
            sum(1 for a in affinities if a < -6.0),
        ),
        ("", ""),
        ("═══ CNN METRICS ═══", ""),
        (
            "Status",
            "NOT AVAILABLE — CNN scoring disabled for this run",
        ),
        (
            "Note",
            "To enable CNN, set --cnn_scoring rescore (not recommended "
            "for metalloprotein covalent docking)",
        ),
        ("", ""),
        ("═══ SANITY CHECK (Tầng 3) ═══", ""),
        ("🟡 Positive Affinity (repulsive)", n_positive),
        (f"🟠 Poor Affinity (≥ {RED_FLAG_AFFINITY_POOR})", n_poor),
        ("✅ Clean (no flags)", done - n_flagged),
        ("", ""),
        ("═══ ARBITRATION THRESHOLDS ═══", ""),
        (
            "Poor Affinity threshold",
            f"≥ {RED_FLAG_AFFINITY_POOR} kcal/mol",
        ),
        (
            "CNN_VS threshold",
            "N/A — CNN disabled for covalent metalloprotein",
        ),
    ]

    _write_header(ws4, 1, ["Metric", "Value"], header_fill_purple)

    for row_idx, (label, value) in enumerate(stats, start=2):
        label_cell = ws4.cell(row=row_idx, column=1, value=label)
        value_cell = ws4.cell(row=row_idx, column=2, value=value)
        label_cell.border = thin_border
        value_cell.border = thin_border

        if "═══" in str(label):
            label_cell.font = Font(bold=True, size=11, color="7030A0")
            label_cell.fill = PatternFill(
                start_color="E8E0F0", end_color="E8E0F0", fill_type="solid"
            )
            value_cell.fill = PatternFill(
                start_color="E8E0F0", end_color="E8E0F0", fill_type="solid"
            )
        else:
            label_cell.font = Font(bold=True)

    _set_col_widths(ws4, {"A": 42, "B": 60})

    excel_path = os.path.join(summary_dir, "docking_summary.xlsx")
    wb.save(excel_path)
    print(f"✔ Excel summary saved to {excel_path}")

    return excel_path


def _generate_csv_fallback_v27(results_sorted: list, summary_dir: str):
    """[v2.7.0] CSV fallback — ranking theo affinity, CNN fields ghi N/A."""
    csv_path = os.path.join(summary_dir, "docking_summary.csv")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        headers = [
            "Energy_Rank",
            "ID",
            "Original_Name",
            "Status",
            "Affinity_kcal",
            "CNNscore",
            "CNN_VS",
            "CNNaffinity",
            "Sanity_Flag",
            "Elapsed_Min",
            "SMILES",
        ]
        writer.writerow(headers)

        for rank, data in enumerate(results_sorted, start=1):
            writer.writerow([
                rank,
                data["lig_id"],
                data["orig_name"],
                data["status"],
                data.get("best_affinity", ""),
                (
                    data.get("best_CNNscore")
                    if data.get("best_CNNscore") is not None
                    else "N/A (CNN off)"
                ),
                (
                    data.get("best_CNN_VS")
                    if data.get("best_CNN_VS") is not None
                    else "N/A (CNN off)"
                ),
                (
                    data.get("best_CNNaffinity")
                    if data.get("best_CNNaffinity") is not None
                    else "N/A (CNN off)"
                ),
                data.get("red_flag", ""),
                data["elapsed_min"],
                data["smiles"],
            ])

    print(f"✔ CSV summary saved to {csv_path}")
    return csv_path


# =========================
# PROGRESS TRACKING
# =========================
def update_progress_csv(ligands: list, summary_dir: str):
    progress_file = os.path.join(summary_dir, "progress.csv")

    with open(progress_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "ID",
                "DIR_NAME",
                "STATUS",
                "ELAPSED_MIN",
                "BEST_AFFINITY",
                "START_TIME",
                "END_TIME",
            ]
        )

        for lig in ligands:
            details = get_status_details(lig["lig_root"])
            writer.writerow(
                [
                    lig["lig_id"],
                    lig["lig_dirname"],
                    details.get("STATUS", STATUS_PENDING),
                    details.get("ELAPSED_MIN", ""),
                    details.get("BEST_AFFINITY", ""),
                    details.get("START_TIME", ""),
                    details.get("END_TIME", ""),
                ]
            )


def print_progress_summary(
    finished: list, failed: list, skipped: list, total: int
):
    done = len(finished) + len(skipped)
    pct = (done / total) * 100 if total > 0 else 0
    print(
        f"\n📊 Progress: {done}/{total} ({pct:.1f}%)"
        f" | ✅ {len(finished)} new"
        f" | ⏭️ {len(skipped)} skipped"
        f" | ❌ {len(failed)} failed"
    )


# =========================
# MAIN PIPELINE
# =========================
def main():
    print("=" * 60)
    print("🧬 GNINA Covalent Flexible Docking Pipeline v2.7.0")
    print("   Metalloprotein Mode (MMP-2 / Zn²⁺)")
    print("   Scoring: Vina empirical only (CNN off)")
    print("   Arbitration: Energy ranking → Sanity Check")
    print("=" * 60)
    print(f"📂 Base dir:  {BASE_DIR}")
    print(f"🔬 GNINA bin: {GNINA_BIN}")
    print(f"🧪 Protein:   {PROTEIN_PATH}")
    print(f"📎 Ref lig:   {REF_LIGAND}")
    print(f"📦 Ligands:   {LIGAND_SDF}")
    print(f"🎯 Flex res:  {FLEX_RESIDUES}")
    print(f"🔗 Covalent:  A:301:ZN ← [OX1;$([O]C=O)]")
    print(f"   Position:  6.739, 10.721, 31.893 Å")
    print(f"🎮 CUDA_VISIBLE_DEVICES = "
          f"{os.environ.get('CUDA_VISIBLE_DEVICES', 'NOT SET')}")
    print(f"🎮 GNINA --device = {GPU_DEVICE}")
    print(f"📊 CNN:       DISABLED (--cnn_scoring none)")
    print(f"📊 Sorting:   energy (minimizedAffinity)")
    print(f"📊 Scoring:   Vina default (NOT vinardo)")
    print(f"🚩 Sanity:    Affinity≥{RED_FLAG_AFFINITY_POOR} or >0 → flagged")
    print(f"⏱️  Timeout:   {GNINA_TIMEOUT_SEC}s "
          f"({GNINA_TIMEOUT_SEC/60:.0f} min/ligand)")
    print("=" * 60)

    if not os.path.isfile(GNINA_BIN):
        print(f"❌ GNINA binary not found: {GNINA_BIN}")
        return
    if not os.access(GNINA_BIN, os.X_OK):
        print(f"❌ GNINA binary not executable: {GNINA_BIN}")
        return

    prepare_root_folders()

    ligands = split_ligands(
        LIGAND_SDF, ligands_root=f"{RESULTS_DIR}/ligands"
    )

    total = len(ligands)
    if total == 0:
        print("❌ No valid ligands found!")
        return

    finished = []
    failed = []
    skipped = []

    print(f"\n🚀 Starting batch covalent docking: {total} ligands\n")
    start_all = time.time()

    for idx, lig in enumerate(ligands, start=1):
        lig_id = lig["lig_id"]
        lig_root = lig["lig_root"]

        status = read_status(lig_root)

        if status == STATUS_DONE:
            print(f"⏭️ [{idx}/{total}] {lig_id} already DONE — skipping")
            skipped.append(lig_id)
            continue

        if status == STATUS_RUNNING:
            print(
                f"⚠️ [{idx}/{total}] {lig_id} was RUNNING (incomplete)"
                " — retrying"
            )

        success = run_gnina(lig, idx, total)

        if success:
            finished.append(lig_id)
        else:
            failed.append(lig_id)

        if idx % 10 == 0:
            update_progress_csv(ligands, f"{RESULTS_DIR}/summary")
            print_progress_summary(finished, failed, skipped, total)

    elapsed_all = (time.time() - start_all) / 60

    summary_dir = f"{RESULTS_DIR}/summary"

    with open(f"{summary_dir}/finished_ligands.txt", "w") as f:
        f.write("\n".join(finished + skipped))

    with open(f"{summary_dir}/failed_ligands.txt", "w") as f:
        f.write("\n".join(failed))

    update_progress_csv(ligands, summary_dir)

    print("\n📊 Generating Excel summary (Covalent Mode — Energy Ranking)...")
    generate_excel_summary(ligands, summary_dir)

    print("\n" + "=" * 60)
    print("📊 FINAL SUMMARY")
    print("=" * 60)
    print(f"⏱️  Total time: {elapsed_all:.2f} min")
    print(f"✅ Completed (new): {len(finished)}")
    print(f"⏭️  Skipped (cached): {len(skipped)}")
    print(f"❌ Failed: {len(failed)}")
    print(f"📁 Results: {RESULTS_DIR}")
    print(f"📊 Excel: {summary_dir}/docking_summary.xlsx")
    print("=" * 60)

    if failed:
        print(
            f"\n⚠️ Failed ligands: {', '.join(failed[:10])}"
            + (
                f"... and {len(failed)-10} more"
                if len(failed) > 10
                else ""
            )
        )


# =============================================================
# Entry point
# =============================================================
if __name__ == "__main__":
    main()
