# ============================================================
# 🧬 GNINA Flexible Docking Pipeline v2.6.2 — Linux + Fish Shell
# ============================================================
# Thay đổi so với v2.6.1:
#   🔴 FIX: Xóa RED_FLAG_CNN_VS_HIGH threshold cho CNN_VS.
#       CNN_VS chỉ dùng để xếp hạng (rank), KHÔNG dùng ngưỡng
#       tuyệt đối. Red Flag giờ chỉ dựa trên thermodynamic
#       sanity (affinity kém hoặc dương/repulsive).
#   Toàn bộ phần còn lại GIỐNG NGUYÊN v2.6.1
# ============================================================

import os
import re
import subprocess
import time
import csv
import shutil
import socket
import traceback
from pathlib import Path
from rdkit import Chem

# =============================================================
# .env loading + LD_LIBRARY_PATH đảm bảo
# =============================================================
NOTEBOOK_DIR = Path.cwd()

_conda_prefix = os.environ.get("CONDA_PREFIX", "")
if _conda_prefix:
    _conda_lib = os.path.join(_conda_prefix, "lib")
    _current_ld = os.environ.get("LD_LIBRARY_PATH", "")
    if _conda_lib not in _current_ld:
        os.environ["LD_LIBRARY_PATH"] = (
            f"{_conda_lib}:{_current_ld}" if _current_ld else _conda_lib
        )
        print(f"🔧 Set LD_LIBRARY_PATH = {os.environ['LD_LIBRARY_PATH']}")

try:
    from dotenv import load_dotenv
    _env_file = NOTEBOOK_DIR / ".env"
    if _env_file.exists():
        load_dotenv(_env_file, override=True)
        print(f"📄 Loaded .env from {_env_file}")
except ImportError:
    print("⚠️ python-dotenv not installed. Using environment variables only.")


def _resolve_path(env_var: str, fallback: str) -> str:
    raw = os.environ.get(env_var, fallback)
    expanded = os.path.expanduser(raw)
    if not os.path.isabs(expanded):
        expanded = str(NOTEBOOK_DIR / expanded)
    return os.path.normpath(expanded)


# =========================
# GLOBAL CONFIG
# =========================
BASE_DIR = _resolve_path("DOCKING_BASE_DIR", str(NOTEBOOK_DIR))
RESULTS_DIR = f"{BASE_DIR}/docking_results"
PROTEIN_PATH = _resolve_path(
    "PROTEIN_PATH",
    f"{BASE_DIR}/data/mmp2_7xjo_ready_for_gnina.pdb",
)
REF_LIGAND = _resolve_path("REF_LIGAND", f"{BASE_DIR}/data/ref_ligand.sdf")
LIGAND_SDF = _resolve_path("LIGAND_SDF", f"{BASE_DIR}/data/ligands_prepared.sdf")
FLEX_RESIDUES = "A:182,A:181,A:215,A:262,A:49"
SEED = "42"
GPU_DEVICE = os.environ.get("GNINA_GPU_DEVICE", "0")

# ================================================================
# Arbitration thresholds
# ================================================================
# [v2.6.2] XÓA RED_FLAG_CNN_VS_HIGH.
#
# Lý do khoa học:
#   CNN_VS (Virtual Screening score) là metric XẾP HẠNG TƯƠNG ĐỐI
#   giữa các ligand trong cùng một chiến dịch sàng lọc. Giới khoa
#   học (bao gồm đội GNINA tại CACHE Challenge #1) KHÔNG thiết lập
#   ngưỡng tuyệt đối cho CNN_VS. Họ dùng nó thuần túy để rank,
#   sau đó trích xuất top-N để phân tích tiếp.
#
#   Ngưỡng cứng 0.70 trước đây nhầm lẫn CNN_VS với CNNscore
#   (xác suất hình học 0–1 của một pose đơn lẻ).
#
# Red Flag giờ chỉ dựa trên thermodynamic sanity:
#   🟡 POSITIVE_AFFINITY: affinity > 0 (repulsive, vô nghĩa vật lý)
#   🟠 POOR_AFFINITY: affinity ≥ threshold (gắn kết quá yếu)
# ================================================================
RED_FLAG_AFFINITY_POOR = -6.5

# =====================================================
# [v2.6.1 FIX 1] GNINA timeout — configurable via .env
# =====================================================
GNINA_TIMEOUT_SEC = int(os.environ.get("GNINA_TIMEOUT_SEC", "3600"))

# GNINA_BIN — auto-detect
_env_bin = os.environ.get("GNINA_BIN", "")
if _env_bin and os.path.isfile(_env_bin):
    GNINA_BIN = _env_bin
elif shutil.which("gnina"):
    GNINA_BIN = shutil.which("gnina")
else:
    GNINA_BIN = f"{BASE_DIR}/bin/gnina"

# Status constants
STATUS_PENDING = "PENDING"
STATUS_RUNNING = "RUNNING"
STATUS_DONE = "DONE"
STATUS_FAILED = "FAILED"


# =========================
# Subprocess env
# =========================
def _get_subprocess_env() -> dict:
    env = os.environ.copy()
    conda_prefix = env.get("CONDA_PREFIX", "")
    if conda_prefix:
        conda_lib = os.path.join(conda_prefix, "lib")
        current_ld = env.get("LD_LIBRARY_PATH", "")
        if conda_lib not in current_ld:
            env["LD_LIBRARY_PATH"] = (
                f"{conda_lib}:{current_ld}" if current_ld else conda_lib
            )
    return env


# =========================
# STATUS MANAGEMENT — giữ nguyên
# =========================
def write_status(lig_root: str, status: str, **kwargs):
    status_file = os.path.join(lig_root, "STATUS.txt")

    if status == STATUS_RUNNING:
        with open(status_file, "w") as f:
            f.write(f"STATUS={status}\n")
            f.write(f"HOST={socket.gethostname()}\n")
            f.write(f"START_TIME={time.strftime('%Y-%m-%d %H:%M:%S')}\n")
    else:
        with open(status_file, "a") as f:
            f.write(f"STATUS={status}\n")
            for key, value in kwargs.items():
                f.write(f"{key.upper()}={value}\n")
            f.write(f"END_TIME={time.strftime('%Y-%m-%d %H:%M:%S')}\n")


def read_status(lig_root: str) -> str:
    status_file = os.path.join(lig_root, "STATUS.txt")

    if not os.path.exists(status_file):
        return STATUS_PENDING

    last_status = STATUS_PENDING
    try:
        with open(status_file, "r") as f:
            for line in f:
                if line.startswith("STATUS="):
                    last_status = line.strip().split("=", 1)[1]
    except Exception:
        pass

    return last_status


def get_status_details(lig_root: str) -> dict:
    status_file = os.path.join(lig_root, "STATUS.txt")
    details = {}

    if os.path.exists(status_file):
        with open(status_file, "r") as f:
            for line in f:
                if "=" in line:
                    key, value = line.strip().split("=", 1)
                    details[key] = value

    return details


# =========================
# UTILS — giữ nguyên
# =========================
def sanitize_name(name: str, max_len: int = 80) -> str:
    if not name:
        return "NA"

    name = name.strip().replace(" ", "_")
    name = re.sub(r'[\/:*?"<>|\\]', "", name)
    name = re.sub(r"_+", "_", name)
    name = name.strip("_")

    return name[:max_len] if name else "NA"


# =========================
# PREPARE ROOT FOLDERS — giữ nguyên
# =========================
def prepare_root_folders():
    dirs = [
        f"{RESULTS_DIR}/protein",
        f"{RESULTS_DIR}/reference",
        f"{RESULTS_DIR}/ligands",
        f"{RESULTS_DIR}/summary",
    ]

    for d in dirs:
        os.makedirs(d, exist_ok=True)

    protein_dest = f"{RESULTS_DIR}/protein/receptor.pdb"
    ref_dest = f"{RESULTS_DIR}/reference/ref_ligand.sdf"

    if not os.path.exists(protein_dest):
        shutil.copy(PROTEIN_PATH, protein_dest)
        print(f"✔ Copied receptor to {protein_dest}")

    if not os.path.exists(ref_dest):
        shutil.copy(REF_LIGAND, ref_dest)
        print(f"✔ Copied reference ligand to {ref_dest}")


# =========================
# SPLIT LIGANDS — giữ nguyên
# =========================
def split_ligands(input_sdf: str, ligands_root: str) -> list:
    os.makedirs(ligands_root, exist_ok=True)

    cleaned_sdf = os.path.join(ligands_root, "_cleaned_input.sdf")
    with open(input_sdf, "rb") as f_in:
        raw = f_in.read()

    if b"\r" in raw:
        raw = raw.replace(b"\r\n", b"\n").replace(b"\r", b"\n")
        with open(cleaned_sdf, "wb") as f_out:
            f_out.write(raw)
        sdf_to_parse = cleaned_sdf
        print(f"🔧 Found \\r in input — normalized line endings")
    else:
        sdf_to_parse = input_sdf
        print(f"✔ Input SDF has clean line endings — no fix needed")

    suppl = Chem.SDMolSupplier(sdf_to_parse, removeHs=False, sanitize=True)
    ligands = []
    mapping = []
    forced_3d_list = []

    for idx, mol in enumerate(suppl, start=1):
        if mol is None:
            print(f"⚠️ Skipping invalid molecule at index {idx}")
            continue

        lig_id = f"LIG_{idx:04d}"

        conf = mol.GetConformer() if mol.GetNumConformers() > 0 else None
        if conf is not None:
            has_z = any(
                abs(conf.GetAtomPosition(i).z) > 0.001
                for i in range(mol.GetNumAtoms())
            )
            if has_z and not conf.Is3D():
                conf.Set3D(True)
                forced_3d_list.append(lig_id)
                print(f"🔧 {lig_id}: forced 3D tag (had Z coordinates)")

        orig_name = mol.GetProp("_Name") if mol.HasProp("_Name") else "NA"
        safe_name = sanitize_name(orig_name)

        lig_dirname = f"{lig_id}__{safe_name}"
        lig_root = os.path.join(ligands_root, lig_dirname)
        input_dir = os.path.join(lig_root, "input")
        os.makedirs(input_dir, exist_ok=True)

        ligand_sdf = os.path.join(input_dir, "ligand.sdf")
        writer = Chem.SDWriter(ligand_sdf)
        writer.SetForceV3000(False)
        writer.write(mol)
        writer.close()

        smiles = Chem.MolToSmiles(mol)
        with open(os.path.join(lig_root, "META.txt"), "w") as f:
            f.write(f"ID={lig_id}\n")
            f.write(f"DIR_NAME={lig_dirname}\n")
            f.write(f"ORIGINAL_NAME={orig_name}\n")
            f.write(f"SMILES={smiles}\n")
            f.write(f"SDF_INDEX={idx}\n")

        ligands.append(
            {
                "lig_id": lig_id,
                "lig_dirname": lig_dirname,
                "lig_root": lig_root,
                "ligand_sdf": ligand_sdf,
                "orig_name": orig_name,
                "smiles": smiles,
            }
        )
        mapping.append((lig_id, lig_dirname, orig_name, smiles))

    mapping_file = os.path.join(ligands_root, "ligand_mapping.csv")
    with open(mapping_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "DIR_NAME", "ORIGINAL_NAME", "SMILES"])
        for row in mapping:
            writer.writerow(row)

    if forced_3d_list:
        forced_log = os.path.join(ligands_root, "forced_3d_log.txt")
        with open(forced_log, "w") as f:
            f.write("# Ligands forced from 2D tag to 3D tag\n")
            f.write(f"# Total: {len(forced_3d_list)}\n")
            for lid in forced_3d_list:
                f.write(f"{lid}\n")
        print(f"📝 Forced 3D log: {forced_log} ({len(forced_3d_list)} ligands)")

    count_3d = 0
    count_2d = 0
    for lig in ligands:
        mol = next(Chem.SDMolSupplier(lig["ligand_sdf"], removeHs=False))
        if mol is not None and mol.GetNumConformers() > 0:
            if mol.GetConformer().Is3D():
                count_3d += 1
            else:
                count_2d += 1
    print(f"✔ Validation: {count_3d} molecules 3D, {count_2d} still 2D")
    if count_2d > 0:
        print(f"⚠️ WARNING: {count_2d} molecules still tagged 2D — check input geometry")

    if os.path.exists(cleaned_sdf):
        os.remove(cleaned_sdf)

    print(f"✔ Split {len(ligands)} ligands")
    print(f"✔ Mapping written to {mapping_file}")

    return ligands


# ================================================================
# SCORE PARSING — Arbitration Protocol (giữ nguyên parse_top_poses)
# ================================================================
def parse_top_poses(sdf_path: str, n: int = 3) -> list:
    """
    Extract top N poses from docked SDF.

    QUAN TRỌNG (Arbitration Protocol — Tầng 1):
    GNINA đã sắp xếp output theo --pose_sort_order CNNscore.
    Hàm này GIỮ NGUYÊN thứ tự từ file (KHÔNG re-sort).
    """
    try:
        suppl = Chem.SDMolSupplier(sdf_path, removeHs=False)
        poses = []

        for pose_idx, mol in enumerate(suppl, start=1):
            if mol is None:
                continue

            pose_data = {"pose_rank": pose_idx}

            score_props = [
                "minimizedAffinity",
                "CNNscore",
                "CNNaffinity",
                "CNN_VS",
                "vinardo",
            ]

            for prop in score_props:
                if mol.HasProp(prop):
                    try:
                        pose_data[prop] = float(mol.GetProp(prop))
                    except ValueError:
                        pose_data[prop] = mol.GetProp(prop)
                else:
                    pose_data[prop] = None

            poses.append(pose_data)

            if len(poses) >= n:
                break

        return poses

    except Exception as e:
        print(f"⚠️ Error parsing scores: {e}")
        return []


# ================================================================
# [v2.6.2] classify_red_flag — XÓA ngưỡng CNN_VS
# ================================================================
def classify_red_flag(affinity) -> str:
    """
    Arbitration Protocol — Tầng 3: Thermodynamic Sanity Check.

    [v2.6.2] Chỉ kiểm tra affinity. CNN_VS KHÔNG có ngưỡng tuyệt đối.

    Lý do khoa học:
      CNN_VS là metric xếp hạng tương đối (relative ranking metric)
      giữa các ligand trong cùng một chiến dịch sàng lọc. Đội ngũ
      tác giả GNINA (CACHE Challenge #1, giải nhất toàn cầu) sử
      dụng CNN_VS thuần túy để rank, KHÔNG áp ngưỡng tuyệt đối.

      Ngưỡng cứng cho CNN_VS gây rủi ro:
        - False negatives trên các túi gắn lạ (novel binding sites)
        - Nhầm lẫn bản chất CNN_VS (ranking) với CNNscore (per-pose
          probability)

    Sanity flags dựa trên nhiệt động học:
      🟡 POSITIVE_AFFINITY: affinity > 0 kcal/mol (repulsive, vô nghĩa
         vật lý — ligand bị đẩy ra khỏi túi gắn)
      🟠 POOR_AFFINITY: affinity ≥ threshold (gắn kết quá yếu, không
         đủ ý nghĩa sinh học tại điều kiện phòng thí nghiệm)
    """
    if affinity is None:
        return ""

    if affinity > 0:
        return "🟡 POSITIVE_AFFINITY"

    if affinity >= RED_FLAG_AFFINITY_POOR:
        return "🟠 POOR_AFFINITY"

    return ""


def get_best_pose_summary(sdf_path: str) -> dict:
    """
    Lấy pose tốt nhất (pose 1 = CNNscore cao nhất) và tính red flag.
    """
    poses = parse_top_poses(sdf_path, n=1)
    if not poses:
        return {
            "CNNscore": None,
            "CNN_VS": None,
            "CNNaffinity": None,
            "minimizedAffinity": None,
            "red_flag": "",
        }

    best = poses[0]
    # [v2.6.2] classify_red_flag chỉ cần affinity
    best["red_flag"] = classify_red_flag(best.get("minimizedAffinity"))
    return best


# ================================================================
# [v2.6.1 FIX 2] parse_best_score — return None thay vì 0.0
# ================================================================
def parse_best_score(sdf_path: str):
    """Extract best minimizedAffinity from docked SDF.

    [v2.6.1] Returns None nếu không parse được.
    Trước đây trả về 0.0 — gây nhầm lẫn vì 0.0 là giá trị
    affinity hợp lệ (dương, repulsive).
    """
    poses = parse_top_poses(sdf_path, n=1)
    if poses and poses[0].get("minimizedAffinity") is not None:
        return poses[0]["minimizedAffinity"]
    return None


# ================================================================
# [v2.6.1 FIX 1+2] RUN GNINA — thêm timeout + xử lý None score
# ================================================================
def run_gnina(ligand_info: dict, idx: int, total: int) -> bool:
    """Run GNINA docking for a single ligand.

    [v2.6.1] Thay đổi so với v2.6:
      - subprocess.run có timeout=GNINA_TIMEOUT_SEC
      - Bắt subprocess.TimeoutExpired riêng
      - parse_best_score trả về None → xử lý đúng trong console + status
    """
    lig_id = ligand_info["lig_id"]
    lig_root = ligand_info["lig_root"]
    ligand_sdf = ligand_info["ligand_sdf"]

    out_dir = os.path.join(lig_root, "output")
    log_dir = os.path.join(lig_root, "logs")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)

    out_lig = os.path.join(out_dir, "docked.sdf")
    out_flex = os.path.join(out_dir, "flex_residues.pdb")
    log_file = os.path.join(log_dir, "gnina.log")
    stderr_file = os.path.join(log_dir, "gnina_stderr.log")
    cmd_file = os.path.join(log_dir, "command.txt")

    write_status(lig_root, STATUS_RUNNING)

    print(f"\n🔄 [{idx}/{total}] Docking {lig_id} ...")
    start = time.time()

    cmd = [
        GNINA_BIN,
        "-r",
        f"{RESULTS_DIR}/protein/receptor.pdb",
        "-l",
        ligand_sdf,
        "--autobox_ligand",
        f"{RESULTS_DIR}/reference/ref_ligand.sdf",
        "--autobox_add",
        "5",
        "--autobox_extend",
        "1",
        "--flexres",
        FLEX_RESIDUES,
        "--num_modes",
        "10",
        "--exhaustiveness",
        "32",
        "--cnn_scoring",
        "rescore",
        "--cnn_empirical_weight",
        "1.0",
        "--pose_sort_order",
        "CNNscore",
        "--device",
        GPU_DEVICE,
        "--seed",
        SEED,
        "--atom_term_data",
        "-o",
        out_lig,
        "--out_flex",
        out_flex,
        "--log",
        log_file,
    ]

    with open(cmd_file, "w") as f:
        f.write(" \\\n    ".join(cmd))

    try:
        with open(stderr_file, "w") as stderr_f:
            subprocess.run(
                cmd,
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=stderr_f,
                text=True,
                env=_get_subprocess_env(),
                # ── [v2.6.1 FIX 1] timeout ──
                timeout=GNINA_TIMEOUT_SEC,
            )

        if not os.path.exists(out_lig) or os.path.getsize(out_lig) == 0:
            raise RuntimeError("GNINA produced empty output SDF")

        elapsed = (time.time() - start) / 60

        # ── [v2.6.1 FIX 2] parse_best_score returns None khi fail ──
        best_score = parse_best_score(out_lig)

        write_status(
            lig_root,
            STATUS_DONE,
            elapsed_min=f"{elapsed:.2f}",
            best_affinity=(
                f"{best_score:.4f}" if best_score is not None else "NA"
            ),
        )

        # Console output — giữ nguyên format, xử lý None
        score_str = f"{best_score:.4f}" if best_score is not None else "N/A"
        print(
            f"✅ [{idx}/{total}] {lig_id} DONE in {elapsed:.2f} min"
            f" (affinity: {score_str})"
        )
        return True

    # ── [v2.6.1 FIX 1] Bắt timeout riêng ──
    except subprocess.TimeoutExpired:
        elapsed = (time.time() - start) / 60
        timeout_min = GNINA_TIMEOUT_SEC / 60
        write_status(
            lig_root,
            STATUS_FAILED,
            elapsed_min=f"{elapsed:.2f}",
            error=f"TIMEOUT after {GNINA_TIMEOUT_SEC}s ({timeout_min:.0f}min)",
        )
        print(
            f"⏰ [{idx}/{total}] {lig_id} TIMEOUT after"
            f" {timeout_min:.0f} min — skipping"
        )
        return False

    except subprocess.CalledProcessError as e:
        elapsed = (time.time() - start) / 60
        write_status(
            lig_root,
            STATUS_FAILED,
            elapsed_min=f"{elapsed:.2f}",
            error=f"GNINA exit code {e.returncode}",
        )
        print(
            f"❌ [{idx}/{total}] {lig_id} FAILED: GNINA exit code {e.returncode}"
        )
        return False

    except Exception as e:
        elapsed = (time.time() - start) / 60
        write_status(
            lig_root,
            STATUS_FAILED,
            elapsed_min=f"{elapsed:.2f}",
            error=str(e),
            traceback=traceback.format_exc().replace("\n", " | "),
        )
        print(f"❌ [{idx}/{total}] {lig_id} FAILED: {e}")
        return False


# ================================================================
# EXCEL SUMMARY — [v2.6.2] cập nhật Red Flag text + xóa CNN_VS threshold
# ================================================================
def generate_excel_summary(ligands: list, summary_dir: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        USE_OPENPYXL = True
    except ImportError:
        print("⚠️ openpyxl not installed, generating CSV instead")
        USE_OPENPYXL = False

    results = []
    for lig in ligands:
        lig_root = lig["lig_root"]
        out_sdf = os.path.join(lig_root, "output", "docked.sdf")

        status = read_status(lig_root)
        details = get_status_details(lig_root)

        top_poses = []
        best_pose = {}
        if status == STATUS_DONE and os.path.exists(out_sdf):
            top_poses = parse_top_poses(out_sdf, n=3)
            best_pose = get_best_pose_summary(out_sdf)

        results.append(
            {
                "lig_id": lig["lig_id"],
                "lig_dirname": lig["lig_dirname"],
                "orig_name": lig["orig_name"],
                "smiles": lig["smiles"],
                "status": status,
                "elapsed_min": details.get("ELAPSED_MIN", ""),
                "top_poses": top_poses,
                "best_CNN_VS": best_pose.get("CNN_VS"),
                "best_CNNscore": best_pose.get("CNNscore"),
                "best_CNNaffinity": best_pose.get("CNNaffinity"),
                "best_affinity": best_pose.get("minimizedAffinity"),
                "red_flag": best_pose.get("red_flag", ""),
            }
        )

    results_by_cnn_vs = sorted(
        results,
        key=lambda x: (
            x["best_CNN_VS"] is None,
            -(x["best_CNN_VS"] or 0),
        ),
    )

    if USE_OPENPYXL:
        _generate_xlsx_v26(results, results_by_cnn_vs, summary_dir)
    else:
        _generate_csv_fallback_v26(results_by_cnn_vs, summary_dir)


def _generate_xlsx_v26(
    results: list, results_by_cnn_vs: list, summary_dir: str
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    header_fill_green = PatternFill(
        start_color="548235", end_color="548235", fill_type="solid"
    )
    header_fill_red = PatternFill(
        start_color="C00000", end_color="C00000", fill_type="solid"
    )
    header_fill_purple = PatternFill(
        start_color="7030A0", end_color="7030A0", fill_type="solid"
    )
    done_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    failed_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    red_flag_fill = PatternFill(
        start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
    )
    caution_fill = PatternFill(
        start_color="FFD93D", end_color="FFD93D", fill_type="solid"
    )
    poor_affinity_fill = PatternFill(
        start_color="FFA94D", end_color="FFA94D", fill_type="solid"
    )
    excellent_fill = PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid"
    )
    good_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def _write_header(ws, row, headers, fill):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    def _set_col_widths(ws, widths: dict):
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    # ── SHEET 1: Intra-Ligand Poses (Tầng 1) ──
    ws1 = wb.active
    ws1.title = "Intra-Ligand_Poses"

    ws1.merge_cells("A1:S1")
    note_cell = ws1.cell(
        row=1,
        column=1,
        value=(
            "TẦNG 1 — INTRA-LIGAND POSE SELECTION: "
            "Poses giữ nguyên thứ tự CNNscore từ GNINA "
            "(pose_sort_order=CNNscore). "
            "Pose 1 = pose tự nhiên nhất về mặt lý sinh."
        ),
    )
    note_cell.font = Font(bold=True, italic=True, size=10, color="2F5496")

    headers_s1 = [
        "#",
        "ID",
        "Original_Name",
        "Status",
        "Time_min",
        "P1_CNNscore",
        "P1_CNN_VS",
        "P1_CNNaffinity",
        "P1_Affinity",
        "P1_Flag",
        "P2_CNNscore",
        "P2_CNN_VS",
        "P2_CNNaffinity",
        "P2_Affinity",
        "P3_CNNscore",
        "P3_CNN_VS",
        "P3_CNNaffinity",
        "P3_Affinity",
        "SMILES",
    ]

    _write_header(ws1, 2, headers_s1, header_fill_blue)
    ws1.freeze_panes = "A3"

    for row_idx, data in enumerate(results_by_cnn_vs, start=3):
        rank = row_idx - 2
        top_poses = data["top_poses"]

        row_data = [
            rank,
            data["lig_id"],
            data["orig_name"],
            data["status"],
            data["elapsed_min"],
        ]

        if len(top_poses) >= 1:
            p = top_poses[0]
            # [v2.6.2] classify_red_flag chỉ cần affinity
            flag = classify_red_flag(p.get("minimizedAffinity"))
            row_data.extend([
                p.get("CNNscore", ""),
                p.get("CNN_VS", ""),
                p.get("CNNaffinity", ""),
                p.get("minimizedAffinity", ""),
                flag,
            ])
        else:
            row_data.extend(["", "", "", "", ""])

        if len(top_poses) >= 2:
            p = top_poses[1]
            row_data.extend([
                p.get("CNNscore", ""),
                p.get("CNN_VS", ""),
                p.get("CNNaffinity", ""),
                p.get("minimizedAffinity", ""),
            ])
        else:
            row_data.extend(["", "", "", ""])

        if len(top_poses) >= 3:
            p = top_poses[2]
            row_data.extend([
                p.get("CNNscore", ""),
                p.get("CNN_VS", ""),
                p.get("CNNaffinity", ""),
                p.get("minimizedAffinity", ""),
            ])
        else:
            row_data.extend(["", "", "", ""])

        row_data.append(data["smiles"])

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_idx == 4:
                if value == STATUS_DONE:
                    cell.fill = done_fill
                elif value == STATUS_FAILED:
                    cell.fill = failed_fill

            # [v2.6.2] Cập nhật flag styling — thêm POOR_AFFINITY
            if col_idx == 10:
                if "POSITIVE" in str(value):
                    cell.fill = caution_fill
                    cell.font = Font(bold=True)
                elif "POOR" in str(value):
                    cell.fill = poor_affinity_fill
                    cell.font = Font(bold=True)

    _set_col_widths(ws1, {
        "A": 5, "B": 12, "C": 35, "D": 10, "E": 9,
        "F": 12, "G": 10, "H": 13, "I": 11, "J": 20,
        "K": 12, "L": 10, "M": 13, "N": 11,
        "O": 12, "P": 10, "Q": 13, "R": 11,
        "S": 60,
    })

    # ── SHEET 2: Inter-Ligand Ranking (Tầng 2) ──
    ws2 = wb.create_sheet(title="Inter-Ligand_Ranking")

    ws2.merge_cells("A1:I1")
    # [v2.6.2] Cập nhật note — CNN_VS chỉ để rank
    note2 = ws2.cell(
        row=1,
        column=1,
        value=(
            "TẦNG 2 — INTER-LIGAND VIRTUAL SCREENING: "
            "Xếp hạng theo CNN_VS (cao = tốt). "
            "CNN_VS là metric xếp hạng tương đối — KHÔNG có ngưỡng "
            "tuyệt đối. Dùng ranking để chọn top-N cho phân tích tiếp. "
            "Cột Flag = Tầng 3 Thermodynamic Sanity Check (chỉ dựa trên affinity)."
        ),
    )
    note2.font = Font(bold=True, italic=True, size=10, color="548235")

    headers_s2 = [
        "VS_Rank",
        "ID",
        "Original_Name",
        "CNN_VS",
        "CNNscore",
        "CNNaffinity",
        "Affinity_kcal",
        "Sanity_Flag",
        "SMILES",
    ]

    _write_header(ws2, 2, headers_s2, header_fill_green)
    ws2.freeze_panes = "A3"

    top_done = [r for r in results_by_cnn_vs if r["status"] == STATUS_DONE][:50]

    for row_idx, data in enumerate(top_done, start=3):
        rank = row_idx - 2

        row_data = [
            rank,
            data["lig_id"],
            data["orig_name"],
            data["best_CNN_VS"],
            data["best_CNNscore"],
            data["best_CNNaffinity"],
            data["best_affinity"],
            data["red_flag"],
            data["smiles"],
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # [v2.6.2] CNN_VS column — NO threshold coloring
            # CNN_VS is a ranking metric, not a threshold metric.
            # Users should interpret rank position, not absolute values.

            if col_idx == 7 and isinstance(value, (int, float)):
                if value > 0:
                    cell.fill = caution_fill
                    cell.font = Font(bold=True, color="C00000")
                elif value >= RED_FLAG_AFFINITY_POOR:
                    cell.fill = PatternFill(
                        start_color="FFF2CC",
                        end_color="FFF2CC",
                        fill_type="solid",
                    )

            # [v2.6.2] Cập nhật flag styling
            if col_idx == 8:
                if "POSITIVE" in str(value):
                    cell.fill = caution_fill
                    cell.font = Font(bold=True)
                elif "POOR" in str(value):
                    cell.fill = poor_affinity_fill
                    cell.font = Font(bold=True)

    _set_col_widths(ws2, {
        "A": 9, "B": 12, "C": 35, "D": 10, "E": 12,
        "F": 13, "G": 13, "H": 22, "I": 60,
    })

    # ── SHEET 3: Red Flags (Tầng 3) ──
    ws3 = wb.create_sheet(title="Sanity_Flags")

    ws3.merge_cells("A1:J1")
    # [v2.6.2] Cập nhật note — không còn CNN_VS threshold
    note3 = ws3.cell(
        row=1,
        column=1,
        value=(
            "TẦNG 3 — THERMODYNAMIC SANITY CHECK: "
            "🟡 POSITIVE_AFFINITY = Affinity > 0 kcal/mol (repulsive, "
            "vô nghĩa vật lý). "
            f"🟠 POOR_AFFINITY = Affinity ≥ {RED_FLAG_AFFINITY_POOR} "
            "kcal/mol (gắn kết quá yếu). "
            "Các ligand này CẦN kiểm tra thủ công bằng visualisation. "
            "CNN_VS KHÔNG được dùng làm tiêu chí gắn cờ (chỉ dùng để rank)."
        ),
    )
    note3.font = Font(bold=True, italic=True, size=10, color="C00000")

    headers_s3 = [
        "#",
        "ID",
        "Original_Name",
        "Flag_Type",
        "CNN_VS",
        "CNNscore",
        "CNNaffinity",
        "Affinity_kcal",
        "Concern",
        "SMILES",
    ]

    _write_header(ws3, 2, headers_s3, header_fill_red)
    ws3.freeze_panes = "A3"

    flagged = [r for r in results_by_cnn_vs if r["red_flag"]]

    if flagged:
        for row_idx, data in enumerate(flagged, start=3):
            flag = data["red_flag"]

            # [v2.6.2] Cập nhật concern messages
            if "POSITIVE" in flag:
                concern = (
                    f"Affinity={data['best_affinity']:.2f} > 0 (repulsive). "
                    "Pose này vi phạm nhiệt động học cơ bản — "
                    "ligand bị đẩy ra khỏi túi gắn."
                )
            elif "POOR" in flag:
                concern = (
                    f"Affinity={data['best_affinity']:.2f} ≥ "
                    f"{RED_FLAG_AFFINITY_POOR} kcal/mol. "
                    "Gắn kết quá yếu — không đủ ý nghĩa sinh học. "
                    "Cần kiểm tra tương tác bằng visualisation."
                )
            else:
                concern = ""

            row_data = [
                row_idx - 2,
                data["lig_id"],
                data["orig_name"],
                flag,
                data["best_CNN_VS"],
                data["best_CNNscore"],
                data["best_CNNaffinity"],
                data["best_affinity"],
                concern,
                data["smiles"],
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # [v2.6.2] Cập nhật styling
                if col_idx == 4:
                    if "POSITIVE" in str(value):
                        cell.fill = caution_fill
                        cell.font = Font(bold=True)
                    elif "POOR" in str(value):
                        cell.fill = poor_affinity_fill
                        cell.font = Font(bold=True)
    else:
        ws3.cell(
            row=3,
            column=1,
            value="✅ Không có ligand nào bị gắn cờ — Tất cả đều pass sanity check.",
        ).font = Font(italic=True, color="548235", size=12)

    _set_col_widths(ws3, {
        "A": 5, "B": 12, "C": 35, "D": 22, "E": 10,
        "F": 12, "G": 13, "H": 13, "I": 60, "J": 60,
    })

    # ── SHEET 4: Statistics ──
    ws4 = wb.create_sheet(title="Statistics")

    total = len(results)
    done = sum(1 for r in results if r["status"] == STATUS_DONE)
    failed = sum(1 for r in results if r["status"] == STATUS_FAILED)
    pending = sum(1 for r in results if r["status"] == STATUS_PENDING)

    done_results = [r for r in results if r["best_CNN_VS"] is not None]
    cnn_vs_values = [r["best_CNN_VS"] for r in done_results]
    affinities = [
        r["best_affinity"]
        for r in done_results
        if r["best_affinity"] is not None
    ]

    # [v2.6.2] Cập nhật flag counting
    n_poor = sum(1 for r in results if "POOR" in r.get("red_flag", ""))
    n_positive = sum(
        1 for r in results if "POSITIVE" in r.get("red_flag", "")
    )
    n_flagged = n_poor + n_positive

    n_timeout = sum(
        1
        for r in results
        if r["status"] == STATUS_FAILED
        and "TIMEOUT" in get_status_details(
            os.path.join(
                RESULTS_DIR, "ligands", r["lig_dirname"]
            )
        ).get("ERROR", "")
    )

    stats = [
        ("═══ PIPELINE OVERVIEW ═══", ""),
        ("Total Ligands", total),
        ("Completed", done),
        ("Failed", failed),
        ("  ↳ of which TIMEOUT", n_timeout),
        ("Pending", pending),
        ("GNINA Timeout Setting", f"{GNINA_TIMEOUT_SEC}s ({GNINA_TIMEOUT_SEC/60:.0f}min)"),
        ("", ""),
        ("═══ CNN_VS DISTRIBUTION (Tầng 2 — Ranking Only) ═══", ""),
        (
            "Best CNN_VS",
            f"{max(cnn_vs_values):.4f}" if cnn_vs_values else "N/A",
        ),
        (
            "Worst CNN_VS",
            f"{min(cnn_vs_values):.4f}" if cnn_vs_values else "N/A",
        ),
        (
            "Mean CNN_VS",
            (
                f"{sum(cnn_vs_values) / len(cnn_vs_values):.4f}"
                if cnn_vs_values
                else "N/A"
            ),
        ),
        # [v2.6.2] Giữ distribution stats nhưng ghi rõ "for reference only"
        (
            "Ligands CNN_VS ≥ 0.80 (ref only)",
            sum(1 for v in cnn_vs_values if v >= 0.80),
        ),
        (
            "Ligands CNN_VS ≥ 0.60 (ref only)",
            sum(1 for v in cnn_vs_values if v >= 0.60),
        ),
        (
            "NOTE",
            "CNN_VS has NO absolute threshold — use ranking position only",
        ),
        ("", ""),
        ("═══ AFFINITY DISTRIBUTION ═══", ""),
        (
            "Best Affinity (kcal/mol)",
            f"{min(affinities):.4f}" if affinities else "N/A",
        ),
        (
            "Worst Affinity (kcal/mol)",
            f"{max(affinities):.4f}" if affinities else "N/A",
        ),
        (
            "Mean Affinity (kcal/mol)",
            (
                f"{sum(affinities) / len(affinities):.4f}"
                if affinities
                else "N/A"
            ),
        ),
        (
            "Ligands Affinity < -8.0",
            sum(1 for a in affinities if a < -8.0),
        ),
        (
            "Ligands Affinity < -7.0",
            sum(1 for a in affinities if a < -7.0),
        ),
        (
            "Ligands Affinity < -6.0",
            sum(1 for a in affinities if a < -6.0),
        ),
        ("", ""),
        ("═══ SANITY CHECK (Tầng 3) ═══", ""),
        ("🟡 Positive Affinity (repulsive)", n_positive),
        (f"🟠 Poor Affinity (≥ {RED_FLAG_AFFINITY_POOR})", n_poor),
        ("✅ Clean (no flags)", done - n_flagged),
        ("", ""),
        ("═══ ARBITRATION THRESHOLDS ═══", ""),
        (
            "Poor Affinity threshold",
            f"≥ {RED_FLAG_AFFINITY_POOR} kcal/mol",
        ),
        (
            "CNN_VS threshold",
            "NONE — ranking metric only (no absolute cutoff)",
        ),
    ]

    _write_header(ws4, 1, ["Metric", "Value"], header_fill_purple)

    for row_idx, (label, value) in enumerate(stats, start=2):
        label_cell = ws4.cell(row=row_idx, column=1, value=label)
        value_cell = ws4.cell(row=row_idx, column=2, value=value)
        label_cell.border = thin_border
        value_cell.border = thin_border

        if "═══" in str(label):
            label_cell.font = Font(bold=True, size=11, color="7030A0")
            label_cell.fill = PatternFill(
                start_color="E8E0F0", end_color="E8E0F0", fill_type="solid"
            )
            value_cell.fill = PatternFill(
                start_color="E8E0F0", end_color="E8E0F0", fill_type="solid"
            )
        else:
            label_cell.font = Font(bold=True)

    _set_col_widths(ws4, {"A": 42, "B": 52})

    excel_path = os.path.join(summary_dir, "docking_summary.xlsx")
    wb.save(excel_path)
    print(f"✔ Excel summary saved to {excel_path}")

    return excel_path


def _generate_csv_fallback_v26(results_sorted: list, summary_dir: str):
    csv_path = os.path.join(summary_dir, "docking_summary.csv")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        headers = [
            "VS_Rank",
            "ID",
            "Original_Name",
            "Status",
            "CNN_VS",
            "CNNscore",
            "CNNaffinity",
            "Affinity_kcal",
            "Sanity_Flag",
            "Elapsed_Min",
            "SMILES",
        ]
        writer.writerow(headers)

        for rank, data in enumerate(results_sorted, start=1):
            writer.writerow([
                rank,
                data["lig_id"],
                data["orig_name"],
                data["status"],
                data.get("best_CNN_VS", ""),
                data.get("best_CNNscore", ""),
                data.get("best_CNNaffinity", ""),
                data.get("best_affinity", ""),
                data.get("red_flag", ""),
                data["elapsed_min"],
                data["smiles"],
            ])

    print(f"✔ CSV summary saved to {csv_path}")
    return csv_path


# =========================
# PROGRESS TRACKING — giữ nguyên
# =========================
def update_progress_csv(ligands: list, summary_dir: str):
    progress_file = os.path.join(summary_dir, "progress.csv")

    with open(progress_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "ID",
                "DIR_NAME",
                "STATUS",
                "ELAPSED_MIN",
                "BEST_AFFINITY",
                "START_TIME",
                "END_TIME",
            ]
        )

        for lig in ligands:
            details = get_status_details(lig["lig_root"])
            writer.writerow(
                [
                    lig["lig_id"],
                    lig["lig_dirname"],
                    details.get("STATUS", STATUS_PENDING),
                    details.get("ELAPSED_MIN", ""),
                    details.get("BEST_AFFINITY", ""),
                    details.get("START_TIME", ""),
                    details.get("END_TIME", ""),
                ]
            )


def print_progress_summary(
    finished: list, failed: list, skipped: list, total: int
):
    done = len(finished) + len(skipped)
    pct = (done / total) * 100 if total > 0 else 0
    print(
        f"\n📊 Progress: {done}/{total} ({pct:.1f}%)"
        f" | ✅ {len(finished)} new"
        f" | ⏭️ {len(skipped)} skipped"
        f" | ❌ {len(failed)} failed"
    )


# =========================
# MAIN PIPELINE
# =========================
def main():
    print("=" * 60)
    print("🧬 GNINA Flexible Docking Pipeline v2.6.2")
    print("   Arbitration Protocol: CNNscore → CNN_VS → Sanity Check")
    print("=" * 60)
    print(f"📂 Base dir:  {BASE_DIR}")
    print(f"🔬 GNINA bin: {GNINA_BIN}")
    print(f"🧪 Protein:   {PROTEIN_PATH}")
    print(f"📎 Ref lig:   {REF_LIGAND}")
    print(f"📦 Ligands:   {LIGAND_SDF}")
    print(f"🎯 Flex res:  {FLEX_RESIDUES}")
    # [v2.6.2] Cập nhật flag display — không còn CNN_VS threshold
    print(f"🚩 Sanity:    Affinity≥{RED_FLAG_AFFINITY_POOR} or >0 → flagged")
    print(f"📊 CNN_VS:    Ranking only (no absolute threshold)")
    print(f"⏱️  Timeout:   {GNINA_TIMEOUT_SEC}s ({GNINA_TIMEOUT_SEC/60:.0f} min/ligand)")
    print("=" * 60)

    if not os.path.isfile(GNINA_BIN):
        print(f"❌ GNINA binary not found: {GNINA_BIN}")
        return
    if not os.access(GNINA_BIN, os.X_OK):
        print(f"❌ GNINA binary not executable: {GNINA_BIN}")
        return

    prepare_root_folders()

    ligands = split_ligands(
        LIGAND_SDF, ligands_root=f"{RESULTS_DIR}/ligands"
    )

    total = len(ligands)
    if total == 0:
        print("❌ No valid ligands found!")
        return

    finished = []
    failed = []
    skipped = []

    print(f"\n🚀 Starting batch docking: {total} ligands\n")
    start_all = time.time()

    for idx, lig in enumerate(ligands, start=1):
        lig_id = lig["lig_id"]
        lig_root = lig["lig_root"]

        status = read_status(lig_root)

        if status == STATUS_DONE:
            print(f"⏭️ [{idx}/{total}] {lig_id} already DONE — skipping")
            skipped.append(lig_id)
            continue

        if status == STATUS_RUNNING:
            print(
                f"⚠️ [{idx}/{total}] {lig_id} was RUNNING (incomplete)"
                " — retrying"
            )

        success = run_gnina(lig, idx, total)

        if success:
            finished.append(lig_id)
        else:
            failed.append(lig_id)

        if idx % 10 == 0:
            update_progress_csv(ligands, f"{RESULTS_DIR}/summary")
            print_progress_summary(finished, failed, skipped, total)

    elapsed_all = (time.time() - start_all) / 60

    summary_dir = f"{RESULTS_DIR}/summary"

    with open(f"{summary_dir}/finished_ligands.txt", "w") as f:
        f.write("\n".join(finished + skipped))

    with open(f"{summary_dir}/failed_ligands.txt", "w") as f:
        f.write("\n".join(failed))

    update_progress_csv(ligands, summary_dir)

    print("\n📊 Generating Excel summary (Arbitration Protocol)...")
    generate_excel_summary(ligands, summary_dir)

    print("\n" + "=" * 60)
    print("📊 FINAL SUMMARY")
    print("=" * 60)
    print(f"⏱️  Total time: {elapsed_all:.2f} min")
    print(f"✅ Completed (new): {len(finished)}")
    print(f"⏭️  Skipped (cached): {len(skipped)}")
    print(f"❌ Failed: {len(failed)}")
    print(f"📁 Results: {RESULTS_DIR}")
    print(f"📊 Excel: {summary_dir}/docking_summary.xlsx")
    print("=" * 60)

    if failed:
        print(
            f"\n⚠️ Failed ligands: {', '.join(failed[:10])}"
            + (
                f"... and {len(failed)-10} more"
                if len(failed) > 10
                else ""
            )
        )


# =============================================================
# Entry point
# =============================================================
if __name__ == "__main__":
    main()
